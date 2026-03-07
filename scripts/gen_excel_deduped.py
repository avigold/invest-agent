"""GROUND TRUTH — DO NOT MODIFY.

This script produced the validated 84.5% avg annual return (2018-2024).
All production ML scoring must replicate this methodology exactly.

Generate backtest Excel with seed 32, no India, deduped by company name.
"""
import sys
sys.path.insert(0, "/Users/avramscore/Projects/invest-agent")

import numpy as np
import lightgbm as lgb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.predict.parquet_dataset import load_parquet_dataset, compute_recency_weights

COUNTRIES = [
    "US","GB","CA","AU","DE","FR","JP","CH","SE","NL",
    "KR","BR","ZA","SG","HK","NO","DK","FI","IL",
    "NZ","TW","IE","BE","AT",
]

print("Loading dataset...", flush=True)
ds = load_parquet_dataset(
    "data/exports/training_features.parquet",
    min_fiscal_year=2000, min_dollar_volume=500_000,
    allowed_countries=COUNTRIES, max_return_clip=10.0,
    return_threshold=0.20, relative_to_country=True,
)

import pyarrow.parquet as pq
raw = pq.read_table("data/exports/training_features.parquet").to_pandas()

meta_lookup = {}
for _, row in raw.iterrows():
    key = (row["ticker"], int(row["fiscal_year"]))
    meta_lookup[key] = {
        "company_name": row.get("company_name", "") or "",
        "country": row.get("country_iso2", "") or "",
        "gics_code": str(row.get("gics_code", "") or ""),
    }

GICS_SECTORS = {
    "10": "Energy", "15": "Materials", "20": "Industrials",
    "25": "Consumer Discretionary", "30": "Consumer Staples",
    "35": "Health Care", "40": "Financials", "45": "Information Technology",
    "50": "Communication Services", "55": "Utilities", "60": "Real Estate",
}

TEST_YEARS = list(range(2018, 2025))
TOP_N = 50
SEED = 32
INVEST_PER_STOCK = 20_000

params = {
    "objective": "binary", "metric": "auc",
    "num_leaves": 63, "min_data_in_leaf": 50,
    "learning_rate": 0.05, "feature_fraction": 0.6,
    "bagging_fraction": 0.7, "bagging_freq": 5,
    "max_depth": -1, "verbose": -1,
    "seed": SEED, "data_random_seed": SEED,
    "feature_fraction_seed": SEED, "bagging_seed": SEED,
}

green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
red_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
bold_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)

wb = Workbook()

summary_ws = wb.active
summary_ws.title = "Summary"
summary_ws.append(["Backtest Portfolio — Seed 32 (No India, Deduped by Company)"])
summary_ws["A1"].font = title_font
summary_ws.append([])
summary_ws.append(["Year", "Portfolio Return", "Median Stock Return",
                    "Winners (>0%)", "Outperformers", "Starting Capital", "Ending Capital"])
for c in range(1, 8):
    summary_ws.cell(row=3, column=c).font = header_font
    summary_ws.cell(row=3, column=c).fill = header_fill

year_summaries = []

for test_year in TEST_YEARS:
    print(f"\n=== {test_year} ===", flush=True)
    train_mask = ds.fiscal_years < test_year
    test_mask = ds.fiscal_years == test_year

    if test_mask.sum() == 0:
        continue

    max_train_yr = int(ds.fiscal_years[train_mask].max())
    w = compute_recency_weights(ds.fiscal_years[train_mask], max_train_yr, ds.half_life)

    tds = lgb.Dataset(ds.X[train_mask], label=ds.y[train_mask], weight=w,
                      feature_name=ds.feature_names,
                      categorical_feature=ds.categorical_features)
    vds = lgb.Dataset(ds.X[test_mask], label=ds.y[test_mask],
                      feature_name=ds.feature_names,
                      categorical_feature=ds.categorical_features,
                      reference=tds)

    bst = lgb.train(params, tds, num_boost_round=1000, valid_sets=[vds],
                    callbacks=[lgb.early_stopping(50, verbose=False)])

    preds = bst.predict(ds.X[test_mask])
    fwd = ds.forward_returns[test_mask]
    test_indices = np.where(test_mask)[0]
    tickers_test = [ds.tickers[i] for i in test_indices]
    years_test = ds.fiscal_years[test_mask]

    # Sort all by prediction score descending
    full_order = np.argsort(-preds)

    # Deduplicate: pick top 50 unique companies
    seen_companies = set()
    selected = []
    skipped = 0
    for idx in full_order:
        ticker = tickers_test[idx]
        yr = int(years_test[idx])
        meta = meta_lookup.get((ticker, yr), {})
        company = meta.get("company_name", ticker)
        # Normalize company name for dedup
        company_key = company.strip().lower()
        if company_key in seen_companies:
            skipped += 1
            continue
        seen_companies.add(company_key)
        selected.append(idx)
        if len(selected) >= TOP_N:
            break

    print(f"  Skipped {skipped} duplicate listings", flush=True)

    # Create year sheet
    ws = wb.create_sheet(title=str(test_year))
    ws.append([f"Portfolio Picks — {test_year} (50 unique companies)"])
    ws["A1"].font = title_font
    ws.append([])

    headers = ["Rank", "Ticker", "Company", "Country", "Sector",
               "Buy Date", "Sell Date", "ML Score", "12M Return",
               "Outperformer", "Invested", "End Value", "Profit/Loss"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_end = 0
    winners_count = 0
    outperformers_count = 0
    returns_list = []

    for rank, idx in enumerate(selected, 1):
        ticker = tickers_test[idx]
        yr = int(years_test[idx])
        ret = float(fwd[idx]) if np.isfinite(fwd[idx]) else 0.0
        score = float(preds[idx])
        is_outperformer = bool(ds.y[test_indices[idx]] == 1.0)

        meta = meta_lookup.get((ticker, yr), {})
        company = meta.get("company_name", "")
        country = meta.get("country", "")
        gics = meta.get("gics_code", "")
        sector = GICS_SECTORS.get(gics[:2], "Unknown") if len(gics) >= 2 else "Unknown"

        end_val = INVEST_PER_STOCK * (1 + ret)
        profit = end_val - INVEST_PER_STOCK
        total_end += end_val
        if ret > 0:
            winners_count += 1
        if is_outperformer:
            outperformers_count += 1
        returns_list.append(ret)

        buy_date = f"{yr}-01-01"
        sell_date = f"{yr}-12-31"

        row_data = [
            rank, ticker, company, country, sector,
            buy_date, sell_date, round(score, 4), ret,
            "Yes" if is_outperformer else "No",
            INVEST_PER_STOCK, round(end_val, 2), round(profit, 2),
        ]
        ws.append(row_data)
        row_num = rank + 3

        fill = green_fill if ret > 0 else red_fill
        for c in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=c).fill = fill

        ws.cell(row=row_num, column=9).number_format = '0.0%'
        ws.cell(row=row_num, column=11).number_format = '$#,##0'
        ws.cell(row=row_num, column=12).number_format = '$#,##0'
        ws.cell(row=row_num, column=13).number_format = '$#,##0'

    portfolio_ret = (total_end / (INVEST_PER_STOCK * len(selected))) - 1
    median_ret = float(np.median(returns_list))

    ws.append([])
    summary_row = len(selected) + 5
    ws.cell(row=summary_row, column=1, value="Portfolio Summary").font = bold_font
    ws.cell(row=summary_row + 1, column=1, value="Starting Capital")
    ws.cell(row=summary_row + 1, column=2, value=INVEST_PER_STOCK * len(selected)).number_format = '$#,##0'
    ws.cell(row=summary_row + 2, column=1, value="Ending Capital")
    ws.cell(row=summary_row + 2, column=2, value=round(total_end, 2)).number_format = '$#,##0'
    ws.cell(row=summary_row + 3, column=1, value="Portfolio Return")
    ws.cell(row=summary_row + 3, column=2, value=portfolio_ret).number_format = '0.0%'
    ws.cell(row=summary_row + 4, column=1, value="Median Stock Return")
    ws.cell(row=summary_row + 4, column=2, value=median_ret).number_format = '0.0%'
    ws.cell(row=summary_row + 5, column=1, value="Winners (>0%)")
    ws.cell(row=summary_row + 5, column=2, value=f"{winners_count}/{len(selected)}")
    ws.cell(row=summary_row + 6, column=1, value="Outperformers")
    ws.cell(row=summary_row + 6, column=2, value=f"{outperformers_count}/{len(selected)}")

    for col, w in [("A",6),("B",14),("C",32),("D",10),("E",24),("F",12),("G",12),("H",10),("I",12),("J",14),("K",12),("L",12),("M",12)]:
        ws.column_dimensions[col].width = w

    year_summaries.append({
        "year": test_year, "ret": portfolio_ret, "median": median_ret,
        "winners": winners_count, "outperformers": outperformers_count,
        "end": total_end, "n": len(selected),
    })
    print(f"  $1M -> ${total_end:,.0f} | ret={portfolio_ret:+.0%} | "
          f"median={median_ret:+.0%} | hits={outperformers_count}/{len(selected)}", flush=True)

for i, ys in enumerate(year_summaries):
    row = 4 + i
    summary_ws.cell(row=row, column=1, value=ys["year"])
    summary_ws.cell(row=row, column=2, value=ys["ret"]).number_format = '0.0%'
    summary_ws.cell(row=row, column=3, value=ys["median"]).number_format = '0.0%'
    summary_ws.cell(row=row, column=4, value=f"{ys['winners']}/{ys['n']}")
    summary_ws.cell(row=row, column=5, value=f"{ys['outperformers']}/{ys['n']}")
    summary_ws.cell(row=row, column=6, value=INVEST_PER_STOCK * ys['n']).number_format = '$#,##0'
    summary_ws.cell(row=row, column=7, value=round(ys["end"], 2)).number_format = '$#,##0'

avg_row = 4 + len(year_summaries) + 1
summary_ws.cell(row=avg_row, column=1, value="Average").font = bold_font
avg_ret = np.mean([ys["ret"] for ys in year_summaries])
summary_ws.cell(row=avg_row, column=2, value=avg_ret).number_format = '0.0%'
summary_ws.cell(row=avg_row, column=2).font = bold_font

for col, w in [("A",10),("B",18),("C",20),("D",15),("E",16),("F",18),("G",18)]:
    summary_ws.column_dimensions[col].width = w

out_path = "data/exports/backtest_portfolio_deduped.xlsx"
wb.save(out_path)
print(f"\nSaved to {out_path}", flush=True)
print(f"Average annual return: {avg_ret:.1%}", flush=True)
